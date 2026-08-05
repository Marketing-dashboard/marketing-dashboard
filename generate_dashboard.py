"""
generate_dashboard.py
---------------------
Reads:
  - Marketing_Performance.xlsx   (quarterly BU data)
  - Enterprise_Comparison.xlsx   (Enterprise model-level Apr vs May)
  - CV_Comparison.xlsx           (CV model-level Apr vs May)

Writes:
  - index.html  (the live dashboard)

Run locally:  python generate_dashboard.py
Runs automatically via GitHub Actions on every push.
"""

import json, math
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl:  pip install openpyxl")


# ── Helpers ───────────────────────────────────────────────────────────────────

def to_float(v):
    """Convert cell value to float; return 0.0 for None/formulas/errors."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        f = float(v)
        return 0.0 if math.isnan(f) or math.isinf(f) else f
    s = str(v).replace("₹","").replace(",","").replace(" ","").strip()
    if s.startswith("=") or not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def jv(v):
    """JS value: None → 'null', number → string."""
    return "null" if v is None else str(v)


# ── 1. Marketing_Performance.xlsx ────────────────────────────────────────────

def read_mp_sheet(wb, sheet_name):
    """
    Reads a sheet where:
      Row 1 = column headers  (Quarter, Spends, Revenue, Leads, Tleads, ...)
      Rows 2+ = data rows     (Q1, 1234, 5678, ...)
    Returns list of dicts keyed by header name.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        entry = {}
        for i, h in enumerate(headers):
            entry[h] = to_float(row[i]) if i < len(row) else 0.0
        entry["_quarter"] = str(row[0]).strip()
        data.append(entry)
    return data

def make_row(q, s, rev, l, t):
    s   = round(s)
    rev = round(rev) if rev else None
    l   = round(l)   if l   else None
    t   = round(t)   if t   else None
    margin = round(1 - s/rev, 4) if rev else None
    cpl    = round(s/l,  3)      if l   else None
    tcpl   = round(s/t,  3)      if t   else None
    return dict(q=q, spends=s, revenue=rev, margin=margin,
                leads=l, tleads=t, cpl=cpl, tcpl=tcpl)

def make_row_no_leads(q, s, rev):
    s   = round(s)
    rev = round(rev) if rev else None
    margin = round(1 - s/rev, 4) if rev else None
    return dict(q=q, spends=s, revenue=rev, margin=margin,
                leads=None, tleads=None, cpl=None, tcpl=None)

def make_row_no_rev(q, s, l, t):
    s=round(s); l=round(l) if l else None; t=round(t) if t else None
    cpl  = round(s/l, 3) if l else None
    tcpl = round(s/t, 3) if t else None
    return dict(q=q, spends=s, revenue=None, margin=None,
                leads=l, tleads=t, cpl=cpl, tcpl=tcpl)

def build_data_js(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Sheet name map: internal key → actual sheet name in Excel
    SHEET_MAP = {
        "newauto":             "New auto classified",
        "enterprise_combined": "Enterprise(Bikes+Cars)",
        "enterprise_cars":     "Enterprise-Cars",
        "enterprise_bikes":    "Enterprise-Bikes",
        "cps":                 "CPS",
        "ncbd":                "NCBD",
        "cv":                  "CV",
    }

    sections = {}

    for key, sheet_name in SHEET_MAP.items():
        rows = read_mp_sheet(wb, sheet_name)
        if not rows:
            print(f"  ⚠ Sheet '{sheet_name}' not found or empty — skipping.")
            continue

        parsed_rows = []
        for r in rows:
            q   = r["_quarter"]
            s   = r.get("Spends", 0)
            rev = r.get("Revenue", 0) or None
            l   = r.get("Leads", 0)   or None
            t   = r.get("Tleads", 0)  or None

            if key == "newauto" or key == "ncbd":
                parsed_rows.append(make_row_no_leads(q, s, rev))
            elif rev:
                parsed_rows.append(make_row(q, s, rev, l or 0, t or 0))
            else:
                parsed_rows.append(make_row_no_rev(q, s, l or 0, t or 0))

        TITLES = {
            "newauto":             "New Auto Classified — Quarter on Quarter Performance",
            "enterprise_combined": "Enterprise — Combined (Cars + Bikes)",
            "enterprise_cars":     "Enterprise — Cars",
            "enterprise_bikes":    "Enterprise — Bikes",
            "cps":                 "CPS — Quarter on Quarter Performance",
            "ncbd":                "NCBD — Quarter on Quarter Performance",
            "cv":                  "CV — Quarter on Quarter Performance",
        }
        KPIS = {
            "newauto":             ["spends","revenue","margin"],
            "enterprise_combined": ["spends","revenue","margin","leads","tleads","cpl","tcpl"],
            "enterprise_cars":     ["spends","revenue","margin","leads","tleads","cpl","tcpl"],
            "enterprise_bikes":    ["spends","revenue","margin","leads","tleads","cpl","tcpl"],
            "cps":                 ["spends","revenue","margin","leads","tleads","cpl","tcpl"],
            "ncbd":                ["spends","revenue","margin"],
            "cv":                  ["spends","revenue","margin","leads","tleads","cpl","tcpl"],
        }

        sections[key] = {
            "title": TITLES[key],
            "kpis":  KPIS[key],
            "rows":  parsed_rows,
        }

    # Build JS string
    js = "const DATA = {\n"
    for key, sec in sections.items():
        js += f'  {key}: {{\n'
        js += f'    title: "{sec["title"]}",\n'
        js += f'    kpis: {json.dumps(sec["kpis"])},\n'
        js += '    rows: [\n'
        for r in sec["rows"]:
            js += (f'      {{ q:"{r["q"]}", spends:{jv(r["spends"])}, '
                   f'revenue:{jv(r["revenue"])}, margin:{jv(r["margin"])}, '
                   f'leads:{jv(r["leads"])}, tleads:{jv(r["tleads"])}, '
                   f'cpl:{jv(r["cpl"])}, tcpl:{jv(r["tcpl"])} }},\n')
        js += '    ]\n  },\n'
    js += '};\n'
    return js


# ── 2. Enterprise_Comparison.xlsx → MODEL_TREE ───────────────────────────────

def build_model_tree(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)

    def read_sheet(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or row[0] is None:
                continue
            brand, model, segment, channel, val_pct, sold_cpl, spends, leads, tleads, revenue, *_ = list(row) + [None]*11
            rows.append({
                "brand": str(brand), "model": str(model), "segment": str(segment),
                "validation": to_float(val_pct),
                "spends": to_float(spends), "leads": int(to_float(leads)),
                "tleads": int(to_float(tleads)), "revenue": to_float(revenue),
            })
        return rows

    def aggregate(rows):
        agg = {}
        for r in rows:
            key = (r["brand"], r["model"], r["segment"])
            if key not in agg:
                agg[key] = {"brand":r["brand"],"model":r["model"],"segment":r["segment"],
                            "spends":0,"leads":0,"tleads":0,"revenue":0,"val_sum":0,"count":0}
            agg[key]["spends"]  += r["spends"]
            agg[key]["leads"]   += r["leads"]
            agg[key]["tleads"]  += r["tleads"]
            agg[key]["revenue"] += r["revenue"]
            agg[key]["val_sum"] += r["validation"]
            agg[key]["count"]   += 1
        out = []
        for key, v in agg.items():
            gcpl   = v["spends"] / v["leads"]   if v["leads"]   else 0
            tcpl   = v["spends"] / v["tleads"]  if v["tleads"]  else 0
            margin = 1 - v["spends"] / v["revenue"] if v["revenue"] else 0
            out.append({"brand":v["brand"],"model":v["model"],"segment":v["segment"],
                        "spends":round(v["spends"]),"leads":v["leads"],"tleads":v["tleads"],
                        "revenue":round(v["revenue"]),"gcpl":round(gcpl,2),"tcpl":round(tcpl,2),
                        "margin":round(margin,4),"validation":round(v["val_sum"]/v["count"],4)})
        return out

    jun_rows = aggregate(read_sheet("Jun'26"))
    jul_rows = aggregate(read_sheet("Jul'26"))

    all_keys = set()
    for r in jun_rows + jul_rows:
        all_keys.add((r["segment"], r["brand"], r["model"]))

    tree = {}
    for seg, brand, model in sorted(all_keys):
        if seg not in tree:          tree[seg] = {}
        if brand not in tree[seg]:   tree[seg][brand] = {}
        jun_r = next((r for r in jun_rows if r["segment"]==seg and r["brand"]==brand and r["model"]==model), None)
        jul_r = next((r for r in jul_rows if r["segment"]==seg and r["brand"]==brand and r["model"]==model), None)
        tree[seg][brand][model] = {"jun": jun_r, "jul": jul_r}

    return f"\nconst MODEL_TREE = {json.dumps(tree)};\n"


# ── 3. CV_Comparison.xlsx → CV_TREE ──────────────────────────────────────────

def build_cv_tree(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Data"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0 = KPI group headers, Row 1 = month sub-headers, Rows 2+ = data
    # Each KPI group has 5 months: Apr(0) May(1) Jun(2) Jul(3) Aug(4)
    # Cols: Brand(0), Model(1), BU(2),
    #   Spends:          3-7  (Apr=3,  May=4,  Jun=5,  Jul=6,  Aug=7)
    #   Leads:           8-12 (Apr=8,  May=9,  Jun=10, Jul=11, Aug=12)
    #   CPL:            13-17 (Apr=13, May=14, Jun=15, Jul=16, Aug=17)
    #   Triggered Leads:18-22 (Apr=18, May=19, Jun=20, Jul=21, Aug=22)
    #   TCPL:           23-27 (Apr=23, May=24, Jun=25, Jul=26, Aug=27)
    #   Margin:         28-32 (Apr=28, May=29, Jun=30, Jul=31, Aug=32)
    #   Validation:     33-37 (Apr=33, May=34, Jun=35, Jul=36, Aug=37)

    tree = {}
    for row in rows[2:]:
        if row[0] is None:
            continue
        brand, model = str(row[0]), str(row[1])
        if brand not in tree:
            tree[brand] = {}

        def make_period(mo):
            # mo: 0=Apr,1=May,2=Jun,3=Jul,4=Aug
            spends = to_float(row[3  + mo])
            leads  = to_float(row[8  + mo])
            tleads = to_float(row[18 + mo])
            return {
                "spends":     spends,
                "leads":      leads,
                "tleads":     tleads,
                "margin":     to_float(row[28 + mo]),
                "validation": to_float(row[33 + mo]),
            } if spends > 0 else None

        tree[brand][model] = {
            "jun": make_period(2),
            "jul": make_period(3),
        }

    return f"\nconst CV_TREE = {json.dumps(tree)};\n"


def detect_current_month(filepath):
    """
    Auto-detect the current (partial) month from Marketing_Performance.xlsx.
    Strategy: the last row in the first data sheet is the current partial month.
    Returns a string like "Jun'26", "Jul'26" etc.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    first_sheet = "Enterprise(Bikes+Cars)"
    if first_sheet not in wb.sheetnames:
        first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]
    last_quarter = None
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None and str(row[0]).strip() not in ("Quarter", ""):
            last_quarter = str(row[0]).strip()
    return last_quarter  # e.g. "Jun'26"


# ── 4. Assemble index.html ────────────────────────────────────────────────────

def main():
    base_dir      = Path(__file__).parent
    template_path = base_dir / "dashboard_template.html"
    output_path   = base_dir / "index.html"

    mp_path  = base_dir / "Marketing_Performance.xlsx"
    ent_path = base_dir / "Enterprise_Comparison.xlsx"
    cv_path  = base_dir / "CV_Comparison.xlsx"

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}\n"
                                "Make sure dashboard_template.html is in the same folder.")

    with open(template_path, encoding="utf-8") as f:
        template = f.read()

    print("📊 Reading Marketing_Performance.xlsx ...")
    if mp_path.exists():
        data_js = build_data_js(mp_path)
        current_month = detect_current_month(mp_path)
        print(f"   ✓ Done  (current partial month: {current_month})")
    else:
        raise FileNotFoundError(f"Marketing_Performance.xlsx not found in {base_dir}")

    print("🚗 Reading Enterprise_Comparison.xlsx ...")
    if ent_path.exists():
        model_tree_js = build_model_tree(ent_path)
        print("   ✓ Done")
    else:
        print("   ⚠ Not found — keeping existing MODEL_TREE in template.")
        model_tree_js = None

    print("🚛 Reading CV_Comparison.xlsx ...")
    if cv_path.exists():
        cv_tree_js = build_cv_tree(cv_path)
        print("   ✓ Done")
    else:
        print("   ⚠ Not found — keeping existing CV_TREE in template.")
        cv_tree_js = None

    # Inject all blocks into template
    current_month_js = f'"{current_month}"' if current_month else '"Jun\'26"'
    html = template.replace("###DATA_PLACEHOLDER###", data_js)
    html = html.replace("###CURRENT_MONTH_PLACEHOLDER###", current_month_js)
    if model_tree_js:
        html = html.replace("###MODEL_TREE_PLACEHOLDER###", model_tree_js)
    if cv_tree_js:
        html = html.replace("###CV_TREE_PLACEHOLDER###", cv_tree_js)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    kb = len(html) // 1024
    print(f"\n✅ index.html generated ({kb} KB)  →  {output_path}")
    print(f"   Current partial month: {current_month}")
    print(f"   Dashboard auto-detects last complete month for summary KPI cards.")


if __name__ == "__main__":
    main()
